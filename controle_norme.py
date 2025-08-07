import re
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from collections import Counter
from tqdm import tqdm
import sys

# === CONFIGURATION API VTOM ===
VTOM_HOST = "https://vtom-admin.carter-cash.com:30002"  # 🔁 À adapter
API_KEY = "B6o5Gfg8PodwKs5u"
HEADERS = {"accept": "application/json", "X-API-KEY": API_KEY}

# === APPELS API VTOM ===
def get_environments():
    return requests.get(f"{VTOM_HOST}/vtom/public/domain/5.0/environments", headers=HEADERS).json()

def get_applications(env):
    return requests.get(f"{VTOM_HOST}/vtom/public/domain/5.0/environments/{env}/applications", headers=HEADERS).json()

def get_jobs(env, app):
    return requests.get(f"{VTOM_HOST}/vtom/public/domain/5.0/environments/{env}/applications/{app}/jobs", headers=HEADERS).json()

def get_job_detail(env, app, job):
    return requests.get(f"{VTOM_HOST}/vtom/public/domain/5.0/environments/{env}/applications/{app}/jobs/{job}", headers=HEADERS).json()

def get_job_links(env, app, job):
    res = requests.get(f"{VTOM_HOST}/vtom/public/domain/5.0/environments/{env}/applications/{app}/jobs/{job}/links", headers=HEADERS)
    return res.json() if res.status_code == 200 else []

def get_job_alarms(env, app, job):
    res = requests.get(f"{VTOM_HOST}/vtom/public/domain/5.0/environments/{env}/applications/{app}/jobs/{job}/alarms", headers=HEADERS)
    return res.text.strip() if res.status_code == 200 else ""


def load_exceptions(filepath):
    apps = set()
    jobs = set()
    with open(filepath, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if ":" in line:
                app, job = line.split(":", 1)
                jobs.add((app.strip(), job.strip()))
            else:
                apps.add(line.strip())
    return apps, jobs


# === GESTION DE LA CRITICITÉ ===
def get_job_criticite(job_name):
    if job_name.startswith("C1"):
        return 1
    elif job_name.startswith("C2"):
        return 2
    elif job_name.startswith("C3"):
        return 3
    return 99

def get_criticite_from_name(name):
    match = re.match(r"(C\d)_", name)
    if match:
        return int(match.group(1)[1])
    return 99

# === CONTRÔLES APPLICATIONS ===
def check_application(env, app, jobs):
    criticite_appli = get_criticite_from_name(app["name"])
    max_job_criticite = min([get_job_criticite(j["name"]) for j in jobs] + [99])
    mode_exec = app.get("execMode", "")

    # Variables pour erreurs
    errors = {}

    # Nomenclature
    if re.match(r"^C\d_[A-Z0-9]{2,3}+_[a-zA-Z0-9_]+$", app["name"]):
        nomenclature = "OK"
    else:
        nomenclature = "KO"
        errors["Nomenclature"] = app["name"]

    # Criticité cohérente
    if criticite_appli <= max_job_criticite:
        crit_coh = "OK"
    else:
        crit_coh = "KO"
        errors["Criticité cohérente"] = f"{criticite_appli}>{max_job_criticite}"

    # Mode exécution
    if mode_exec == "Job":
        mode_exec_val = "OK"
    else:
        mode_exec_val = "KO"
        errors["Mode exécution"] = mode_exec

    result = {
        "ENV": env,
        "APPLICATION": app["name"],
        "Nomenclature": nomenclature,
        "Criticité cohérente": crit_coh,
        "Mode exécution": mode_exec_val,
    }

    result["STATUT APPLICATION"] = "OK" if all(v == "OK" for k, v in result.items() if k not in ["ENV", "APPLICATION"]) else "KO"
    result["DETAIL_ERREUR"] = ", ".join([f"{k}: {v}" for k, v in errors.items()])
    return result

# === CONTRÔLES JOBS ===
def check_job(env, app_name, job):
    job_name = job.get("name", "")
    detail = get_job_detail(env, app_name, job_name)
    links = get_job_links(env, app_name, job_name)
    alarms = get_job_alarms(env, app_name, job_name)
    contexts = job.get("contexts", [])
    instruction = detail.get("instruction", "").strip()
    criticite = get_job_criticite(job_name)

    errors = {}

    # Nomenclature
    if re.match(r"^C\d_[A-Z0-9]{2,3}_[A-Z0-9]{2,3}(?:_[A-Z0-9]{2,3})?_[A-Za-z0-9]+$", job_name):
        nomenclature = "OK"
    else:
        nomenclature = "KO"
        errors["Nomenclature"] = job_name

    # Consigne obligatoire
    if criticite in [1, 2]:
        if instruction:
            consigne = "OK"
        else:
            consigne = "KO"
            errors["Consigne obligatoire"] = "(vide)"
    else:
        consigne = "N/A"

    # Lien Mandatory C1 (nouvelle règle)
    lien_mandatory = "N/A"
    if links:
        # On cherche les liens entrants (source = job amont)
        lien_mandatory = "N/A"
        for link in links:
            try:
                # link["source"] = "environnement/application/job"
                _, _, source_job = link["target"].split("/")
                source_crit = get_job_criticite(source_job)
                if source_crit == 1:
                    if link.get("type") == "Mandatory":
                        lien_mandatory = "OK"
                    else:
                        lien_mandatory = "KO"
                        errors["Lien Mandatory C1"] = f"{source_job} ({link.get('type', 'inconnu')})"
                        break  # On s'arrête au premier KO
            except Exception as e:
                continue

    # Dépendances Criticité
    dep_crit = "N/A"
    if links:
        mandatory_in_links = [l for l in links if l.get("type") == "Mandatory"]
        if mandatory_in_links:
            dep_ko = None
            for link in mandatory_in_links:
                try:
                    _, _, pred_job = link["target"].split("/")
                    pred_crit = get_job_criticite(pred_job)
                    if pred_crit > criticite:
                        dep_ko = pred_job
                        break
                except:
                    continue
            if dep_ko:
                dep_crit = "KO"
                errors["Dépendances Criticité"] = dep_ko
            else:
                dep_crit = "OK"

    # Variable SSIS
    if "_SSI_" in job_name:
        if not any("VT_workDir" in str(c) for c in contexts):
            ssis = "KO"
            errors["Variable SSIS"] = "Absente"
        else:
            ssis = "OK"
    else:
        ssis = "N/A"

    # Alarme KO V3
    if "ITM_JOB_KO_V3" in alarms:
        alarme_ko = "OK"
    else:
        alarme_ko = "KO"
        errors["Alarme KO V3"] = alarms

    # Max Run Time C1
    if criticite == 1:
        if any("ITM_JOB_MRT_" in a for a in alarms.split()):
            mrt = "OK"
        else:
            mrt = "KO"
            errors["Max Run Time C1"] = "Non trouvé"
    else:
        mrt = "N/A"

    result = {
        "ENV": env,
        "APPLICATION": app_name,
        "JOB": job_name,
        "Nomenclature": nomenclature,
        "Consigne obligatoire": consigne,
        "Lien Mandatory C1": lien_mandatory,
        "Dépendances Criticité": dep_crit,
        "Variable SSIS": ssis,
        "Alarme KO V3": alarme_ko,
        "Max Run Time C1": mrt,
    }

    result["STATUT JOB"] = "OK" if all(v in ["OK", "N/A"] for k, v in result.items() if k not in ["ENV", "APPLICATION", "JOB"]) else "KO"
    result["DETAIL_ERREUR"] = ", ".join([f"{k}: {v}" for k, v in errors.items()])
    return result

# === EXÉCUTION PRINCIPALE ===
if __name__ == "__main__":
    # Récupère le fichier d'exceptions en argument
    if len(sys.argv) < 2:
        print("Usage: python controle_norme.py exceptions.txt")
        sys.exit(1)
    exceptions_file = sys.argv[1]
    except_apps, except_jobs = load_exceptions(exceptions_file)

    appli_results = []
    job_results = []

    environments = get_environments()
    for env in environments:
        env_name = env["name"]
        applications = get_applications(env_name)
        for app in tqdm(applications, desc=f"[{env_name}] Applications", unit="appli"):
            app_name = app["name"]
            if app_name in except_apps:
                continue  # Ignore toute l'application
            jobs = get_jobs(env_name, app_name)
            appli_results.append(check_application(env_name, app, jobs))
            for job in tqdm(jobs, desc=f"  [{app_name}] Jobs", unit="job", leave=False):
                job_name = job.get("name", "")
                if (app_name, job_name) in except_jobs:
                    continue  # Ignore ce job précis
                job_results.append(check_job(env_name, app_name, job))

    # === EXPORT EXCEL ===
    wb = Workbook()

    # Onglet Applications
    ws_app = wb.active
    ws_app.title = "Check Appli"
    df_app = pd.DataFrame(appli_results)
    for r in dataframe_to_rows(df_app, index=False, header=True):
        ws_app.append(r)

    # Onglet Jobs
    ws_job = wb.create_sheet("Check Jobs")
    df_job = pd.DataFrame(job_results)
    for r in dataframe_to_rows(df_job, index=False, header=True):
        ws_job.append(r)

    # Mise en forme rouge sur KO
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for ws in [ws_app, ws_job]:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value == "KO":
                    cell.fill = red_fill

    # Sauvegarde avec timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"Controle_norme_VTOM_global_{timestamp}.xlsx"
    excel_path = f"D:/Extract_Norme/{excel_filename}"
    wb.save(excel_path)

    # === RÉSUMÉ CONSOLE ===
    def analyze_errors(results, statut_key):
        error_counter = Counter()
        for r in results:
            for key, val in r.items():
                if key not in ["ENV", "APPLICATION", "JOB", "STATUT JOB", "STATUT APPLICATION"]:
                    if val == "KO":
                        error_counter[key] += 1
        return error_counter.most_common()

    nb_applis = len(appli_results)
    nb_applis_ok = sum(1 for r in appli_results if r["STATUT APPLICATION"] == "OK")
    nb_applis_ko = nb_applis - nb_applis_ok

    nb_jobs = len(job_results)
    nb_jobs_ok = sum(1 for r in job_results if r["STATUT JOB"] == "OK")
    nb_jobs_ko = nb_jobs - nb_jobs_ok

    print(f"🔍 Contrôle terminé")
    print(f"🧾 Applications analysées : {nb_applis} | ✅ OK : {nb_applis_ok} | ❌ KO : {nb_applis_ko}")
    print(f"🧾 Jobs analysés        : {nb_jobs} | ✅ OK : {nb_jobs_ok} | ❌ KO : {nb_jobs_ko}")

    print("\n📊 Erreurs fréquentes - Applications :")
    for key, count in analyze_errors(appli_results, "STATUT APPLICATION"):
        print(f" - {key}: {count} KO")

    print("\n📊 Erreurs fréquentes - Jobs :")
    for key, count in analyze_errors(job_results, "STATUT JOB"):
        print(f" - {key}: {count} KO")

    print(f"\n📄 Fichier généré : {excel_path}")


